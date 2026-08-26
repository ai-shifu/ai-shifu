"""Shared helpers for the LLM cache overcharge compensation scripts."""

from __future__ import annotations

import csv
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

DEFAULT_INPUT_PATH = "cache_overcharge.csv"
DEFAULT_SHEET_NAME = "\u6700\u7ec8\u8865\u53d1\u6e05\u5355"
DEFAULT_CAMPAIGN_ID = "llm-cache-overcharge-20260826"
DEFAULT_OPERATOR_USER_BID = "cache-overcharge-compensation-script"
USER_BID_HEADER = "\u8d26\u6237ID"
IDENTIFY_HEADER = "\u8d26\u53f7/\u624b\u673a\u53f7"
NICKNAME_HEADER = "\u6635\u79f0"
SCENES_HEADER = "\u6d89\u53ca\u573a\u666f"
AMOUNT_HEADER = "\u5efa\u8bae\u8865\u53d1\u79ef\u5206"


def ensure_api_root_on_path() -> None:
    """Ensure imports work when the script is executed by file path."""
    api_root = Path(__file__).resolve().parents[1]
    if str(api_root) not in sys.path:
        sys.path.insert(0, str(api_root))


ensure_api_root_on_path()

from flaskr.service.billing.primitives import quantize_credit_amount  # noqa: E402


@dataclass(slots=True, frozen=True)
class CompensationInputRow:
    """One row from the compensation reference table."""

    user_bid: str
    identify: str
    nickname: str
    scenes: str
    amount: Decimal
    row_number: int


class JsonEncoder(json.JSONEncoder):
    """JSON encoder for script output."""

    def default(self, value: object) -> object:
        """Serialize non-standard JSON values used by script payloads."""
        if isinstance(value, Decimal):
            return format(value, "f")
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return super().default(value)


def dump_json(payload: dict[str, object]) -> None:
    """Print stable JSON output."""
    print(json.dumps(payload, ensure_ascii=False, indent=2, cls=JsonEncoder))


def load_reference_rows(
    input_path: str,
    *,
    sheet_name: str,
) -> list[CompensationInputRow]:
    """Load compensation reference rows from xlsx or csv."""
    path = Path(input_path).expanduser()
    if not path.exists():
        message = f"Input file does not exist: {path}"
        raise FileNotFoundError(message)
    if path.suffix.lower() == ".csv":
        return _load_csv_rows(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(path, sheet_name=sheet_name)
    message = "Input file must be .xlsx, .xlsm, or .csv"
    raise ValueError(message)


def filter_rows_by_user_bid(
    rows: list[CompensationInputRow],
    user_bids: list[str],
) -> list[CompensationInputRow]:
    """Filter rows to the requested user bids."""
    normalized = {str(user_bid or "").strip() for user_bid in user_bids}
    normalized.discard("")
    if not normalized:
        return rows
    return [row for row in rows if row.user_bid in normalized]


def row_to_payload(row: CompensationInputRow) -> dict[str, object]:
    """Serialize one input row for script output."""
    return {
        "row_number": row.row_number,
        "user_bid": row.user_bid,
        "identify": row.identify,
        "nickname": row.nickname,
        "scenes": row.scenes,
        "amount": row.amount,
    }


def _load_csv_rows(path: Path) -> list[CompensationInputRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[CompensationInputRow] = []
        for row_number, row in enumerate(reader, start=2):
            parsed = _parse_mapping_row(row, row_number=row_number)
            if parsed is not None:
                rows.append(parsed)
        return rows


def _load_xlsx_rows(path: Path, *, sheet_name: str) -> list[CompensationInputRow]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        message = f"Sheet not found: {sheet_name}"
        raise ValueError(message)
    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []
    header_map = {
        str(header or "").strip(): index for index, header in enumerate(headers)
    }
    rows: list[CompensationInputRow] = []
    for row_number, values in enumerate(iterator, start=2):
        mapping = {
            header: values[index] if index < len(values) else None
            for header, index in header_map.items()
        }
        parsed = _parse_mapping_row(mapping, row_number=row_number)
        if parsed is not None:
            rows.append(parsed)
    return rows


def _parse_mapping_row(
    row: dict[str, Any],
    *,
    row_number: int,
) -> CompensationInputRow | None:
    user_bid = _normalize_cell(row.get(USER_BID_HEADER))
    if not user_bid:
        return None
    return CompensationInputRow(
        user_bid=user_bid,
        identify=_normalize_cell(row.get(IDENTIFY_HEADER)),
        nickname=_normalize_cell(row.get(NICKNAME_HEADER)),
        scenes=_normalize_cell(row.get(SCENES_HEADER)),
        amount=_parse_amount(row.get(AMOUNT_HEADER)),
        row_number=row_number,
    )


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_amount(value: object) -> Decimal:
    try:
        parsed = Decimal(str(value or "0").strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(0)
    return quantize_credit_amount(parsed)
