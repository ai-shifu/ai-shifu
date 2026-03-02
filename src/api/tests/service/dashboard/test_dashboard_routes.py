from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from flaskr.dao import db
from flaskr.service.learn.models import LearnGeneratedBlock, LearnProgressRecord
from flaskr.service.order.consts import (
    LEARN_STATUS_COMPLETED,
    LEARN_STATUS_IN_PROGRESS,
    LEARN_STATUS_NOT_STARTED,
)
from flaskr.service.order.models import Order
from flaskr.service.shifu.consts import (
    BLOCK_TYPE_MDANSWER_VALUE,
    BLOCK_TYPE_MDASK_VALUE,
    BLOCK_TYPE_MDINTERACTION_VALUE,
    UNIT_TYPE_VALUE_GUEST,
    UNIT_TYPE_VALUE_NORMAL,
)
from flaskr.service.shifu.models import (
    AiCourseAuth,
    LogPublishedStruct,
    PublishedShifu,
    PublishedOutlineItem,
    ShifuUserArchive,
)
from flaskr.service.shifu.shifu_history_manager import HistoryItem


@dataclass
class _OutlineSeed:
    bid: str
    title: str
    type: int = UNIT_TYPE_VALUE_NORMAL
    hidden: int = 0
    parent_bid: str = ""


@pytest.mark.usefixtures("app")
class TestDashboardRoutes:
    def _mock_request_user(self, monkeypatch, *, user_id: str = "teacher-1"):
        dummy_user = SimpleNamespace(
            user_id=user_id,
            language="en-US",
            is_creator=True,
        )
        monkeypatch.setattr(
            "flaskr.route.user.validate_user",
            lambda _app, _token: dummy_user,
            raising=False,
        )

    def _allow_permission(self, monkeypatch, *, allowed: bool):
        monkeypatch.setattr(
            "flaskr.service.dashboard.funcs.shifu_permission_verification",
            lambda _app, _user_id, _shifu_bid, _perm: allowed,
            raising=False,
        )

    def _seed_published_struct(self, *, shifu_bid: str, outlines: list[_OutlineSeed]):
        now = datetime.utcnow()
        rows: list[PublishedOutlineItem] = []
        for idx, outline in enumerate(outlines):
            rows.append(
                PublishedOutlineItem(
                    outline_item_bid=outline.bid,
                    shifu_bid=shifu_bid,
                    title=outline.title,
                    type=outline.type,
                    hidden=outline.hidden,
                    parent_bid=outline.parent_bid,
                    position=str(idx),
                    content="",
                    deleted=0,
                    created_at=now,
                    updated_at=now,
                )
            )
        db.session.add_all(rows)
        db.session.flush()

        struct = HistoryItem(
            bid=shifu_bid,
            id=0,
            type="shifu",
            children=[
                HistoryItem(
                    bid=row.outline_item_bid,
                    id=row.id,
                    type="outline",
                    children=[],
                )
                for row in rows
            ],
        )
        db.session.add(
            LogPublishedStruct(
                struct_bid="struct-" + shifu_bid,
                shifu_bid=shifu_bid,
                struct=struct.to_json(),
                deleted=0,
                created_at=now,
            )
        )
        db.session.commit()
        return rows

    def _seed_dashboard_course(
        self,
        *,
        shifu_bid: str,
        title: str,
        user_id: str = "teacher-1",
    ) -> None:
        now = datetime.utcnow()
        db.session.add(
            PublishedShifu(
                shifu_bid=shifu_bid,
                title=title,
                description="",
                avatar_res_bid="",
                llm="",
                llm_temperature=0,
                llm_system_prompt="",
                ask_enabled_status=0,
                ask_llm="",
                ask_llm_temperature=0,
                ask_llm_system_prompt="",
                price=0,
                deleted=0,
                created_at=now,
                created_user_bid=user_id,
                updated_at=now,
                updated_user_bid=user_id,
            )
        )

    def _seed_shared_course_auth(
        self,
        *,
        shifu_bid: str,
        user_id: str = "teacher-1",
        auth_type: str = '["view"]',
        status: int = 1,
    ) -> None:
        now = datetime.utcnow()
        db.session.add(
            AiCourseAuth(
                course_auth_id=f"auth-{user_id}-{shifu_bid}",
                course_id=shifu_bid,
                user_id=user_id,
                auth_type=auth_type,
                status=status,
                created_at=now,
                updated_at=now,
            )
        )

    def test_entry_summary_uses_shared_and_owned_courses(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)

        now = datetime(2025, 1, 15, 10, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(shifu_bid="course-a", title="Course A")
            self._seed_dashboard_course(
                shifu_bid="course-b",
                title="Course B",
                user_id="another-teacher",
            )
            self._seed_shared_course_auth(shifu_bid="course-b")
            self._seed_shared_course_auth(shifu_bid="course-b")
            db.session.add(
                ShifuUserArchive(
                    shifu_bid="course-b",
                    user_bid="teacher-1",
                    archived=1,
                    archived_at=now,
                    created_at=now,
                    updated_at=now,
                )
            )

            db.session.add_all(
                [
                    LearnProgressRecord(
                        progress_record_bid="entry-progress-a-1",
                        shifu_bid="course-a",
                        outline_item_bid="outline-1",
                        user_bid="learner-1",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                    LearnProgressRecord(
                        progress_record_bid="entry-progress-a-2",
                        shifu_bid="course-a",
                        outline_item_bid="outline-1",
                        user_bid="learner-2",
                        status=LEARN_STATUS_NOT_STARTED,
                        block_position=0,
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                    LearnProgressRecord(
                        progress_record_bid="entry-progress-b-1",
                        shifu_bid="course-b",
                        outline_item_bid="outline-2",
                        user_bid="learner-3",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                ]
            )

            db.session.add_all(
                [
                    Order(
                        order_bid="order-a-1",
                        shifu_bid="course-a",
                        user_bid="learner-1",
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                    Order(
                        order_bid="order-a-2",
                        shifu_bid="course-a",
                        user_bid="learner-2",
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                    Order(
                        order_bid="order-b-1",
                        shifu_bid="course-b",
                        user_bid="learner-3",
                        deleted=0,
                        created_at=now,
                        updated_at=now,
                    ),
                ]
            )

            db.session.add_all(
                [
                    LearnGeneratedBlock(
                        generated_block_bid="entry-ask-a-1",
                        progress_record_bid="entry-progress-a-1",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="outline-1",
                        shifu_bid="course-a",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q1",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=now,
                        updated_at=now,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="entry-ask-a-2",
                        progress_record_bid="entry-progress-a-2",
                        user_bid="learner-2",
                        block_bid="",
                        outline_item_bid="outline-1",
                        shifu_bid="course-a",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q2",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=now,
                        updated_at=now,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="entry-ask-b-1",
                        progress_record_bid="entry-progress-b-1",
                        user_bid="learner-3",
                        block_bid="",
                        outline_item_bid="outline-2",
                        shifu_bid="course-b",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="QB",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=now,
                        updated_at=now,
                    ),
                ]
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/entry?page_index=1&page_size=20")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 2
        assert payload["data"]["summary"]["learner_count"] == 3
        assert payload["data"]["summary"]["order_count"] == 3
        assert payload["data"]["summary"]["generation_count"] == 3
        assert payload["data"]["total"] == 2
        assert len(payload["data"]["items"]) == 2
        assert {item["shifu_bid"] for item in payload["data"]["items"]} == {
            "course-a",
            "course-b",
        }

    def test_entry_keyword_and_date_range_filters(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)

        in_range = datetime(2025, 1, 10, 9, 0, 0)
        out_of_range = datetime(2024, 12, 20, 9, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(shifu_bid="course-alg", title="Algebra 101")
            self._seed_dashboard_course(shifu_bid="course-bio", title="Biology 101")

            db.session.add_all(
                [
                    LearnProgressRecord(
                        progress_record_bid="entry-filter-progress-alg",
                        shifu_bid="course-alg",
                        outline_item_bid="outline-1",
                        user_bid="learner-a",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnProgressRecord(
                        progress_record_bid="entry-filter-progress-bio",
                        shifu_bid="course-bio",
                        outline_item_bid="outline-2",
                        user_bid="learner-b",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnProgressRecord(
                        progress_record_bid="entry-filter-progress-created-out",
                        shifu_bid="course-alg",
                        outline_item_bid="outline-3",
                        user_bid="learner-c",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=out_of_range,
                        updated_at=in_range,
                    ),
                ]
            )

            db.session.add_all(
                [
                    Order(
                        order_bid="entry-filter-order-in",
                        shifu_bid="course-alg",
                        user_bid="learner-a",
                        deleted=0,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    Order(
                        order_bid="entry-filter-order-out",
                        shifu_bid="course-alg",
                        user_bid="learner-a",
                        deleted=0,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                ]
            )

            db.session.add_all(
                [
                    LearnGeneratedBlock(
                        generated_block_bid="entry-filter-ask-in",
                        progress_record_bid="entry-filter-progress-alg",
                        user_bid="learner-a",
                        block_bid="",
                        outline_item_bid="outline-1",
                        shifu_bid="course-alg",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="In range",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="entry-filter-ask-out",
                        progress_record_bid="entry-filter-progress-alg",
                        user_bid="learner-a",
                        block_bid="",
                        outline_item_bid="outline-1",
                        shifu_bid="course-alg",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Out range",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                ]
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/entry"
            "?keyword=alG"
            "&start_date=2025-01-01"
            "&end_date=2025-01-31"
            "&page_index=1&page_size=20"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 1
        assert payload["data"]["summary"]["learner_count"] == 1
        assert payload["data"]["summary"]["order_count"] == 1
        assert payload["data"]["summary"]["generation_count"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-alg"
        assert payload["data"]["items"][0]["learner_count"] == 1
        assert payload["data"]["items"][0]["order_count"] == 1
        assert payload["data"]["items"][0]["generation_count"] == 1

    def test_entry_course_count_respects_date_filter(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)

        in_range = datetime(2025, 2, 10, 9, 0, 0)
        out_of_range = datetime(2024, 11, 20, 9, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(shifu_bid="course-a", title="Course A")
            self._seed_dashboard_course(shifu_bid="course-b", title="Course B")

            db.session.add(
                Order(
                    order_bid="entry-date-order-a",
                    shifu_bid="course-a",
                    user_bid="learner-a",
                    deleted=0,
                    created_at=in_range,
                    updated_at=in_range,
                )
            )
            db.session.add(
                Order(
                    order_bid="entry-date-order-b",
                    shifu_bid="course-b",
                    user_bid="learner-b",
                    deleted=0,
                    created_at=out_of_range,
                    updated_at=out_of_range,
                )
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/entry"
            "?start_date=2025-02-01"
            "&end_date=2025-02-28"
            "&page_index=1&page_size=20"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 1
        assert payload["data"]["total"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-a"

    def test_entry_order_only_user_not_counted_as_learner(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)

        now = datetime(2025, 2, 10, 9, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(shifu_bid="course-order", title="Order Course")
            db.session.add(
                Order(
                    order_bid="order-only-1",
                    shifu_bid="course-order",
                    user_bid="imported-user",
                    deleted=0,
                    created_at=now,
                    updated_at=now,
                )
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/entry"
            "?start_date=2025-02-01"
            "&end_date=2025-02-28"
            "&page_index=1&page_size=20"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["learner_count"] == 0
        assert payload["data"]["summary"]["order_count"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-order"
        assert payload["data"]["items"][0]["learner_count"] == 0
        assert payload["data"]["items"][0]["order_count"] == 1

    def test_entry_shared_course_auth_requires_view_only_and_status_1(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)

        with app.app_context():
            self._seed_dashboard_course(
                shifu_bid="course-owned",
                title="Owned Course",
            )
            self._seed_dashboard_course(
                shifu_bid="course-view",
                title="Shared View",
                user_id="teacher-2",
            )
            self._seed_dashboard_course(
                shifu_bid="course-edit",
                title="Shared Edit",
                user_id="teacher-2",
            )
            self._seed_dashboard_course(
                shifu_bid="course-publish",
                title="Shared Publish",
                user_id="teacher-2",
            )
            self._seed_dashboard_course(
                shifu_bid="course-mixed",
                title="Shared Mixed",
                user_id="teacher-2",
            )
            self._seed_dashboard_course(
                shifu_bid="course-disabled",
                title="Shared Disabled",
                user_id="teacher-2",
            )
            self._seed_shared_course_auth(
                shifu_bid="course-view",
                auth_type='["view"]',
                status=1,
            )
            self._seed_shared_course_auth(
                shifu_bid="course-edit",
                auth_type='["edit"]',
                status=1,
            )
            self._seed_shared_course_auth(
                shifu_bid="course-publish",
                auth_type='["publish"]',
                status=1,
            )
            self._seed_shared_course_auth(
                shifu_bid="course-mixed",
                auth_type='["view","edit"]',
                status=1,
            )
            self._seed_shared_course_auth(
                shifu_bid="course-disabled",
                auth_type='["view"]',
                status=0,
            )
            # Duplicate valid auth rows should still dedupe by course_id.
            self._seed_shared_course_auth(
                shifu_bid="course-view",
                auth_type='["view"]',
                status=1,
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/entry?page_index=1&page_size=20")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 2
        assert payload["data"]["total"] == 2
        assert {item["shifu_bid"] for item in payload["data"]["items"]} == {
            "course-owned",
            "course-view",
        }

    def test_entry_excludes_stale_shared_courses(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        monkeypatch.setattr(
            "flaskr.service.dashboard.funcs.get_dynamic_config",
            lambda _key, default=None: default,
            raising=False,
        )

        with app.app_context():
            self._seed_dashboard_course(
                shifu_bid="course-live",
                title="Live Course",
            )
            self._seed_shared_course_auth(
                shifu_bid="course-stale",
                auth_type='["view"]',
                status=1,
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/entry?page_index=1&page_size=20")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 1
        assert payload["data"]["total"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-live"

    def test_entry_excludes_demo_courses(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        monkeypatch.setattr(
            "flaskr.service.dashboard.funcs.get_dynamic_config",
            lambda key, default=None: "course-demo"
            if key == "DEMO_SHIFU_BID"
            else default,
            raising=False,
        )

        with app.app_context():
            self._seed_dashboard_course(
                shifu_bid="course-demo",
                title="Demo Course",
            )
            self._seed_dashboard_course(
                shifu_bid="course-live",
                title="Live Course",
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/entry?page_index=1&page_size=20")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 1
        assert payload["data"]["total"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-live"

    def test_entry_excludes_builtin_demo_titles_when_config_missing(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        monkeypatch.setattr(
            "flaskr.service.dashboard.funcs.get_dynamic_config",
            lambda _key, default=None: default,
            raising=False,
        )

        with app.app_context():
            self._seed_dashboard_course(
                shifu_bid="e867343eaab44488ad792ec54d8b82b5",
                title="AI 师傅教学引导",
                user_id="system",
            )
            self._seed_dashboard_course(
                shifu_bid="b5d7844387e940ed9480a6f945a6db6a",
                title="AI-Shifu Creation Guide",
                user_id="system",
            )
            self._seed_dashboard_course(
                shifu_bid="course-live",
                title="Live Course",
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/entry?page_index=1&page_size=20")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["summary"]["course_count"] == 1
        assert payload["data"]["total"] == 1
        assert payload["data"]["items"][0]["shifu_bid"] == "course-live"

    def test_outlines_requires_permission(self, monkeypatch, test_client):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=False)

        resp = test_client.get("/api/dashboard/shifus/course-1/outlines")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] != 0

    def test_outlines_returns_struct_order(self, monkeypatch, test_client, app):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-1",
                outlines=[
                    _OutlineSeed(bid="o-1", title="Outline 1"),
                    _OutlineSeed(bid="o-2", title="Outline 2"),
                ],
            )

        resp = test_client.get("/api/dashboard/shifus/course-1/outlines")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert [item["outline_item_bid"] for item in payload["data"]] == ["o-1", "o-2"]

    def test_overview_excludes_hidden_and_uses_latest_progress(
        self, monkeypatch, test_client, app
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        in_range = datetime(2025, 6, 15, 10, 0, 0)
        out_of_range = datetime(2024, 12, 15, 10, 0, 0)
        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-2",
                outlines=[
                    _OutlineSeed(
                        bid="r-1", title="Required 1", type=UNIT_TYPE_VALUE_NORMAL
                    ),
                    _OutlineSeed(
                        bid="hidden-1",
                        title="Hidden required",
                        type=UNIT_TYPE_VALUE_NORMAL,
                        hidden=1,
                    ),
                    _OutlineSeed(
                        bid="guest-1",
                        title="Guest",
                        type=UNIT_TYPE_VALUE_GUEST,
                    ),
                ],
            )

            # Learner A: latest record is completed (older is not started)
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-a-1",
                    shifu_bid="course-2",
                    outline_item_bid="r-1",
                    user_bid="learner-a",
                    status=LEARN_STATUS_NOT_STARTED,
                    block_position=0,
                    deleted=0,
                    created_at=in_range - timedelta(days=2),
                    updated_at=in_range - timedelta(days=2),
                )
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-a-2",
                    shifu_bid="course-2",
                    outline_item_bid="r-1",
                    user_bid="learner-a",
                    status=LEARN_STATUS_COMPLETED,
                    block_position=10,
                    deleted=0,
                    created_at=in_range - timedelta(days=1),
                    updated_at=in_range - timedelta(days=1),
                )
            )

            # Learner B: has only guest progress, still counts as a learner
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-b-1",
                    shifu_bid="course-2",
                    outline_item_bid="guest-1",
                    user_bid="learner-b",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=1,
                    deleted=0,
                    created_at=in_range - timedelta(days=1),
                    updated_at=in_range - timedelta(days=1),
                )
            )

            # Learner C: updated in range but created out of range, should be excluded
            # from learner_count because learner_count uses created_at range.
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-c-1",
                    shifu_bid="course-2",
                    outline_item_bid="r-1",
                    user_bid="learner-c",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=1,
                    deleted=0,
                    created_at=out_of_range,
                    updated_at=in_range,
                )
            )

            db.session.add_all(
                [
                    Order(
                        order_bid="overview-order-in",
                        shifu_bid="course-2",
                        user_bid="learner-a",
                        deleted=0,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    Order(
                        order_bid="overview-order-out",
                        shifu_bid="course-2",
                        user_bid="learner-a",
                        deleted=0,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="overview-gen-in",
                        progress_record_bid="attend-a-2",
                        user_bid="learner-a",
                        block_bid="",
                        outline_item_bid="r-1",
                        shifu_bid="course-2",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q in",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="overview-gen-out",
                        progress_record_bid="attend-a-2",
                        user_bid="learner-a",
                        block_bid="",
                        outline_item_bid="r-1",
                        shifu_bid="course-2",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q out",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                ]
            )
            db.session.add(
                LearnGeneratedBlock(
                    generated_block_bid="overview-follow-up-in",
                    progress_record_bid="attend-a-2",
                    user_bid="learner-a",
                    block_bid="",
                    outline_item_bid="r-1",
                    shifu_bid="course-2",
                    type=BLOCK_TYPE_MDASK_VALUE,
                    role=2,
                    generated_content="Trend in",
                    position=1,
                    block_content_conf="",
                    liked=0,
                    deleted=0,
                    status=1,
                    created_at=in_range + timedelta(hours=1),
                    updated_at=in_range + timedelta(hours=1),
                )
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-2/overview?start_date=2025-01-01&end_date=2025-12-31"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["kpis"]["required_outline_total"] == 1
        assert payload["data"]["kpis"]["learner_count"] == 2
        assert payload["data"]["kpis"]["order_count"] == 1
        assert payload["data"]["kpis"]["generation_count"] == 2
        assert payload["data"]["kpis"]["last_active_at"] != ""
        assert payload["data"]["kpis"]["follow_up_ask_total"] == 2
        assert payload["data"]["kpis"]["completion_count"] == 1
        assert sum(point["value"] for point in payload["data"]["follow_up_trend"]) == 2
        chapter_distribution = {
            point["label"]: point["value"]
            for point in payload["data"]["follow_up_chapter_distribution"]
        }
        assert chapter_distribution.get("Required 1") == 2

    def test_generation_count_consistent_between_entry_and_overview(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        in_range = datetime(2025, 2, 10, 9, 0, 0)
        out_of_range = datetime(2024, 12, 10, 9, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(
                shifu_bid="course-consistent", title="Consistent"
            )
            self._seed_published_struct(
                shifu_bid="course-consistent",
                outlines=[_OutlineSeed(bid="lesson-1", title="Lesson 1")],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="consistent-progress",
                    shifu_bid="course-consistent",
                    outline_item_bid="lesson-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=0,
                    deleted=0,
                    created_at=in_range,
                    updated_at=in_range,
                )
            )
            db.session.add_all(
                [
                    LearnGeneratedBlock(
                        generated_block_bid="consistent-gen-1",
                        progress_record_bid="consistent-progress",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-consistent",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q1",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="consistent-gen-2",
                        progress_record_bid="consistent-progress",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-consistent",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q2",
                        position=2,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=in_range + timedelta(hours=1),
                        updated_at=in_range + timedelta(hours=1),
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="consistent-gen-out",
                        progress_record_bid="consistent-progress",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-consistent",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Q-out",
                        position=3,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                ]
            )
            db.session.commit()

        query = "start_date=2025-02-01&end_date=2025-02-28"
        entry_resp = test_client.get(
            f"/api/dashboard/entry?{query}&page_index=1&page_size=20"
        )
        entry_payload = entry_resp.get_json(force=True)
        overview_resp = test_client.get(
            f"/api/dashboard/shifus/course-consistent/overview?{query}"
        )
        overview_payload = overview_resp.get_json(force=True)

        assert entry_resp.status_code == 200
        assert overview_resp.status_code == 200
        assert entry_payload["code"] == 0
        assert overview_payload["code"] == 0
        assert entry_payload["data"]["items"][0]["generation_count"] == 2
        assert overview_payload["data"]["kpis"]["generation_count"] == 2

    def test_overview_without_date_range_uses_full_history_and_returns_shifu_name(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        old_time = datetime(2023, 1, 1, 9, 0, 0)
        with app.app_context():
            self._seed_dashboard_course(shifu_bid="course-all", title="Course All")
            self._seed_published_struct(
                shifu_bid="course-all",
                outlines=[_OutlineSeed(bid="lesson-1", title="Lesson 1")],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="all-progress-1",
                    shifu_bid="course-all",
                    outline_item_bid="lesson-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=0,
                    deleted=0,
                    created_at=old_time,
                    updated_at=old_time,
                )
            )
            db.session.add(
                LearnGeneratedBlock(
                    generated_block_bid="all-ask-1",
                    progress_record_bid="all-progress-1",
                    user_bid="learner-1",
                    block_bid="",
                    outline_item_bid="lesson-1",
                    shifu_bid="course-all",
                    type=BLOCK_TYPE_MDASK_VALUE,
                    role=2,
                    generated_content="old question",
                    position=1,
                    block_content_conf="",
                    liked=0,
                    deleted=0,
                    status=1,
                    created_at=old_time,
                    updated_at=old_time,
                )
            )
            db.session.commit()

        resp = test_client.get("/api/dashboard/shifus/course-all/overview")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["shifu_name"] == "Course All"
        assert payload["data"]["start_date"] == ""
        assert payload["data"]["end_date"] == ""
        assert payload["data"]["kpis"]["learner_count"] == 1
        assert payload["data"]["kpis"]["follow_up_ask_total"] == 1

    def test_learners_progress_and_detail_only_use_leaf_outlines(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        now = datetime(2025, 2, 1, 9, 0, 0)
        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-leaf",
                outlines=[
                    _OutlineSeed(bid="chapter-1", title="Chapter 1"),
                    _OutlineSeed(
                        bid="lesson-1",
                        title="Lesson 1",
                        parent_bid="chapter-1",
                    ),
                ],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="leaf-progress-1",
                    shifu_bid="course-leaf",
                    outline_item_bid="lesson-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_COMPLETED,
                    block_position=10,
                    deleted=0,
                    created_at=now,
                    updated_at=now,
                )
            )
            db.session.commit()

        learners_resp = test_client.get(
            "/api/dashboard/shifus/course-leaf/learners?page_index=1&page_size=20"
        )
        learners_payload = learners_resp.get_json(force=True)
        assert learners_resp.status_code == 200
        assert learners_payload["code"] == 0
        assert learners_payload["data"]["items"][0]["required_outline_total"] == 1
        assert learners_payload["data"]["items"][0]["completed_outline_count"] == 1
        assert learners_payload["data"]["items"][0]["progress_percent"] == 1.0

        detail_resp = test_client.get(
            "/api/dashboard/shifus/course-leaf/learners/learner-1"
        )
        detail_payload = detail_resp.get_json(force=True)
        assert detail_resp.status_code == 200
        assert detail_payload["code"] == 0
        assert [
            item["outline_item_bid"] for item in detail_payload["data"]["outlines"]
        ] == ["lesson-1"]

    def test_learners_pagination(self, monkeypatch, test_client, app):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-3",
                outlines=[_OutlineSeed(bid="r-1", title="Required 1")],
            )
            for idx in range(3):
                db.session.add(
                    LearnProgressRecord(
                        progress_record_bid=f"attend-{idx}",
                        shifu_bid="course-3",
                        outline_item_bid="r-1",
                        user_bid=f"learner-{idx}",
                        status=LEARN_STATUS_NOT_STARTED,
                        block_position=0,
                        deleted=0,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow(),
                    )
                )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-3/learners?page_index=1&page_size=2"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["total"] == 3
        assert payload["data"]["page_count"] == 2
        assert len(payload["data"]["items"]) == 2

    def test_followups_pairs_answer(self, monkeypatch, test_client, app):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        now = datetime.utcnow()
        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-4",
                outlines=[_OutlineSeed(bid="r-1", title="Required 1")],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-1",
                    shifu_bid="course-4",
                    outline_item_bid="r-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=0,
                    deleted=0,
                    created_at=now,
                    updated_at=now,
                )
            )

            # Ask then answer
            db.session.add(
                LearnGeneratedBlock(
                    generated_block_bid="ask-1",
                    progress_record_bid="attend-1",
                    user_bid="learner-1",
                    block_bid="",
                    outline_item_bid="r-1",
                    shifu_bid="course-4",
                    type=BLOCK_TYPE_MDASK_VALUE,
                    role=2,
                    generated_content="Q1",
                    position=1,
                    block_content_conf="",
                    liked=0,
                    deleted=0,
                    status=1,
                    created_at=now,
                    updated_at=now,
                )
            )
            db.session.add(
                LearnGeneratedBlock(
                    generated_block_bid="answer-1",
                    progress_record_bid="attend-1",
                    user_bid="learner-1",
                    block_bid="",
                    outline_item_bid="r-1",
                    shifu_bid="course-4",
                    type=BLOCK_TYPE_MDANSWER_VALUE,
                    role=1,
                    generated_content="A1",
                    position=1,
                    block_content_conf="",
                    liked=0,
                    deleted=0,
                    status=1,
                    created_at=now + timedelta(seconds=5),
                    updated_at=now + timedelta(seconds=5),
                )
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-4/learners/learner-1/followups"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["total"] == 1
        assert payload["data"]["items"][0]["question"] == "Q1"
        assert payload["data"]["items"][0]["answer"] == "A1"

    def test_followups_without_date_range_use_full_history(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        old_time = datetime(2023, 1, 1, 10, 0, 0)
        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-followup-all",
                outlines=[_OutlineSeed(bid="lesson-1", title="Lesson 1")],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="followup-all-progress",
                    shifu_bid="course-followup-all",
                    outline_item_bid="lesson-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=0,
                    deleted=0,
                    created_at=old_time,
                    updated_at=old_time,
                )
            )
            db.session.add(
                LearnGeneratedBlock(
                    generated_block_bid="followup-all-ask",
                    progress_record_bid="followup-all-progress",
                    user_bid="learner-1",
                    block_bid="",
                    outline_item_bid="lesson-1",
                    shifu_bid="course-followup-all",
                    type=BLOCK_TYPE_MDASK_VALUE,
                    role=2,
                    generated_content="old ask",
                    position=1,
                    block_content_conf="",
                    liked=0,
                    deleted=0,
                    status=1,
                    created_at=old_time,
                    updated_at=old_time,
                )
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-followup-all/learners/learner-1/followups"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["total"] == 1

    def test_learner_detail_followups_respects_date_range(
        self,
        monkeypatch,
        test_client,
        app,
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        in_range = datetime(2025, 2, 10, 9, 0, 0)
        out_of_range = datetime(2024, 12, 20, 9, 0, 0)
        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-5",
                outlines=[_OutlineSeed(bid="r-1", title="Required 1")],
            )
            db.session.add(
                LearnProgressRecord(
                    progress_record_bid="attend-5",
                    shifu_bid="course-5",
                    outline_item_bid="r-1",
                    user_bid="learner-1",
                    status=LEARN_STATUS_IN_PROGRESS,
                    block_position=0,
                    deleted=0,
                    created_at=in_range,
                    updated_at=in_range,
                )
            )
            db.session.add_all(
                [
                    LearnGeneratedBlock(
                        generated_block_bid="detail-ask-in",
                        progress_record_bid="attend-5",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="r-1",
                        shifu_bid="course-5",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="In range ask",
                        position=1,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=in_range,
                        updated_at=in_range,
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="detail-ask-out",
                        progress_record_bid="attend-5",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="r-1",
                        shifu_bid="course-5",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="Out range ask",
                        position=2,
                        block_content_conf="",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=out_of_range,
                        updated_at=out_of_range,
                    ),
                ]
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-5/learners/learner-1"
            "?start_date=2025-02-01&end_date=2025-02-28"
        )
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] == 0
        assert payload["data"]["followups"]["total_ask_count"] == 1
        assert (
            payload["data"]["followups"]["by_outline"][0]["outline_item_bid"] == "r-1"
        )
        assert payload["data"]["followups"]["by_outline"][0]["ask_count"] == 1

    def test_dashboard_export_requires_permission(self, monkeypatch, test_client):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=False)

        resp = test_client.get("/api/dashboard/shifus/course-export/export")
        payload = resp.get_json(force=True)

        assert resp.status_code == 200
        assert payload["code"] != 0

    def test_dashboard_export_returns_expected_excel(
        self, monkeypatch, test_client, app
    ):
        self._mock_request_user(monkeypatch)
        self._allow_permission(monkeypatch, allowed=True)

        with app.app_context():
            self._seed_published_struct(
                shifu_bid="course-export",
                outlines=[
                    _OutlineSeed(bid="chapter-1", title="Chapter 1"),
                    _OutlineSeed(
                        bid="lesson-1",
                        title="Lesson 1",
                        parent_bid="chapter-1",
                    ),
                    _OutlineSeed(
                        bid="lesson-2",
                        title="Lesson 2",
                        parent_bid="chapter-1",
                    ),
                ],
            )

            db.session.add_all(
                [
                    LearnProgressRecord(
                        progress_record_bid="progress-1",
                        shifu_bid="course-export",
                        outline_item_bid="lesson-1",
                        user_bid="learner-1",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=datetime(2025, 1, 31, 9, 0, 0),
                        updated_at=datetime(2025, 2, 10, 9, 0, 0),
                    ),
                    LearnProgressRecord(
                        progress_record_bid="progress-2",
                        shifu_bid="course-export",
                        outline_item_bid="lesson-2",
                        user_bid="learner-2",
                        status=LEARN_STATUS_IN_PROGRESS,
                        block_position=0,
                        deleted=0,
                        created_at=datetime(2025, 2, 2, 10, 0, 0),
                        updated_at=datetime(2025, 2, 10, 10, 0, 0),
                    ),
                ]
            )

            db.session.add_all(
                [
                    LearnGeneratedBlock(
                        generated_block_bid="row-interaction-1",
                        progress_record_bid="progress-1",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-export",
                        type=BLOCK_TYPE_MDINTERACTION_VALUE,
                        role=2,
                        generated_content="interaction input 1",
                        position=2,
                        block_content_conf="interaction block 1",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=datetime(2025, 2, 10, 9, 0, 0),
                        updated_at=datetime(2025, 2, 10, 9, 0, 0),
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="row-ask-1",
                        progress_record_bid="progress-1",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-export",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="ask content 1",
                        position=2,
                        block_content_conf="ask fallback 1",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=datetime(2025, 2, 10, 9, 1, 0),
                        updated_at=datetime(2025, 2, 10, 9, 1, 0),
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="row-ask-2",
                        progress_record_bid="progress-2",
                        user_bid="learner-2",
                        block_bid="",
                        outline_item_bid="lesson-2",
                        shifu_bid="course-export",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="ask content 2",
                        position=1,
                        block_content_conf="ask fallback 2",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=datetime(2025, 2, 10, 10, 0, 0),
                        updated_at=datetime(2025, 2, 10, 10, 0, 0),
                    ),
                    LearnGeneratedBlock(
                        generated_block_bid="row-out-of-range",
                        progress_record_bid="progress-1",
                        user_bid="learner-1",
                        block_bid="",
                        outline_item_bid="lesson-1",
                        shifu_bid="course-export",
                        type=BLOCK_TYPE_MDASK_VALUE,
                        role=2,
                        generated_content="out of range ask",
                        position=2,
                        block_content_conf="out of range",
                        liked=0,
                        deleted=0,
                        status=1,
                        created_at=datetime(2024, 12, 20, 9, 0, 0),
                        updated_at=datetime(2024, 12, 20, 9, 0, 0),
                    ),
                ]
            )
            db.session.commit()

        resp = test_client.get(
            "/api/dashboard/shifus/course-export/export"
            "?start_date=2025-02-01&end_date=2025-02-28"
        )

        assert resp.status_code == 200
        assert (
            resp.headers.get("Content-Type")
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        workbook = load_workbook(filename=BytesIO(resp.data))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        assert rows[0] == (
            "学生ID",
            "章节",
            "课时",
            "对话项内容",
            "进入时间",
            "学生输入内容",
        )
        assert len(rows) == 4
        assert rows[1] == (
            "learner-1",
            "Chapter 1",
            "Lesson 1",
            "interaction block 1",
            "2025-01-31 09:00:00",
            "interaction input 1",
        )
        assert rows[2] == (
            "learner-1",
            "Chapter 1",
            "Lesson 1",
            "interaction block 1",
            "2025-01-31 09:00:00",
            "ask content 1",
        )
        assert rows[3] == (
            "learner-2",
            "Chapter 1",
            "Lesson 2",
            "ask fallback 2",
            "2025-02-02 10:00:00",
            "ask content 2",
        )
